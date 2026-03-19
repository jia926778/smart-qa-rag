"""基于 SQLite 的结构化数据存储模块，支持 Text-to-SQL 功能。

当上传 Excel/CSV 文件时，其表格内容自动加载到本地 SQLite 数据库
（每个工作表/文件对应一个表）。Text-to-SQL 智能体随后可以从
自然语言问题生成 SQL 查询，并从结构化数据中返回精确结果。
"""

from __future__ import annotations

import os
import re
import sqlite3
from typing import Any, Dict, List, Optional, Tuple

from app.config import Settings
from app.utils.logger import get_logger

logger = get_logger(__name__)


def _sanitize_table_name(name: str) -> str:
    """Make a safe SQLite table name."""
    name = re.sub(r"[^\w]", "_", name)
    name = re.sub(r"_+", "_", name).strip("_").lower()
    if not name or name[0].isdigit():
        name = "t_" + name
    return name[:64]


def _sanitize_column_name(name: str) -> str:
    """Make a safe column name."""
    name = re.sub(r"[^\w]", "_", str(name))
    name = re.sub(r"_+", "_", name).strip("_").lower()
    if not name:
        name = "col"
    if name[0].isdigit():
        name = "c_" + name
    return name[:64]


class SQLStore:
    """Manage a SQLite database for structured data from uploaded files.

    Each collection gets its own SQLite database file under
    ``{CHROMA_PERSIST_DIR}/../sql_stores/``.
    """

    def __init__(self, settings: Settings) -> None:
        self._settings = settings
        self._store_dir = os.path.join(
            os.path.dirname(settings.CHROMA_PERSIST_DIR), "sql_stores"
        )
        os.makedirs(self._store_dir, exist_ok=True)

    def _db_path(self, collection_name: str) -> str:
        return os.path.join(self._store_dir, f"{collection_name}.db")

    def _connect(self, collection_name: str) -> sqlite3.Connection:
        return sqlite3.connect(self._db_path(collection_name))

    # -----------------------------------------------------------------
    # Ingest: load tabular data into SQL tables
    # -----------------------------------------------------------------

    def ingest_table(
        self,
        collection_name: str,
        table_name: str,
        columns: List[str],
        rows: List[List[Any]],
        source_file: str,
    ) -> Dict[str, Any]:
        """Insert a table into SQLite.  Replaces if table already exists."""
        safe_name = _sanitize_table_name(table_name)
        safe_cols = [_sanitize_column_name(c) for c in columns]

        # Deduplicate column names
        seen = {}
        deduped = []
        for c in safe_cols:
            if c in seen:
                seen[c] += 1
                deduped.append(f"{c}_{seen[c]}")
            else:
                seen[c] = 0
                deduped.append(c)
        safe_cols = deduped

        conn = self._connect(collection_name)
        try:
            cur = conn.cursor()

            # Drop existing table
            cur.execute(f'DROP TABLE IF EXISTS "{safe_name}"')

            # Create table
            col_defs = ", ".join(f'"{c}" TEXT' for c in safe_cols)
            cur.execute(f'CREATE TABLE "{safe_name}" ({col_defs})')

            # Insert rows
            placeholders = ", ".join(["?"] * len(safe_cols))
            insert_sql = f'INSERT INTO "{safe_name}" VALUES ({placeholders})'
            for row in rows:
                # Pad or trim row to match columns
                padded = list(row[:len(safe_cols)])
                while len(padded) < len(safe_cols):
                    padded.append(None)
                # Convert all to string
                padded = [str(v) if v is not None else None for v in padded]
                cur.execute(insert_sql, padded)

            # Store metadata in a registry table
            cur.execute("""
                CREATE TABLE IF NOT EXISTS _table_registry (
                    table_name TEXT PRIMARY KEY,
                    source_file TEXT,
                    columns TEXT,
                    row_count INTEGER
                )
            """)
            cur.execute("""
                INSERT OR REPLACE INTO _table_registry
                VALUES (?, ?, ?, ?)
            """, (safe_name, source_file, ", ".join(safe_cols), len(rows)))

            conn.commit()

            logger.info(
                "SQL ingest: table '%s' (%d cols, %d rows) from '%s' into collection '%s'",
                safe_name, len(safe_cols), len(rows), source_file, collection_name,
            )

            return {
                "table_name": safe_name,
                "columns": safe_cols,
                "rows_count": len(rows),
                "source_file": source_file,
            }
        finally:
            conn.close()

    def ingest_from_excel(
        self,
        collection_name: str,
        file_path: str,
        source_file: str,
    ) -> List[Dict[str, Any]]:
        """Load all sheets from an Excel file into SQL tables."""
        results: List[Dict[str, Any]] = []
        try:
            import openpyxl  # type: ignore
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(list(row))
                if len(rows) < 2:
                    continue  # Need at least header + 1 data row
                columns = [str(c) if c else f"col_{i}" for i, c in enumerate(rows[0])]
                data_rows = rows[1:]
                table_name = f"{_sanitize_table_name(source_file)}_{_sanitize_table_name(sheet_name)}"
                info = self.ingest_table(collection_name, table_name, columns, data_rows, source_file)
                results.append(info)
            wb.close()
        except ImportError:
            logger.warning("openpyxl not available for SQL ingest")
        except Exception as exc:
            logger.warning("Excel SQL ingest failed: %s", exc)
        return results

    def ingest_from_csv(
        self,
        collection_name: str,
        file_path: str,
        source_file: str,
    ) -> Optional[Dict[str, Any]]:
        """Load a CSV file into a SQL table."""
        import csv
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                rows = list(reader)
            if len(rows) < 2:
                return None
            columns = [str(c) if c else f"col_{i}" for i, c in enumerate(rows[0])]
            data_rows = rows[1:]
            table_name = _sanitize_table_name(source_file)
            return self.ingest_table(collection_name, table_name, columns, data_rows, source_file)
        except Exception as exc:
            logger.warning("CSV SQL ingest failed: %s", exc)
            return None

    # -----------------------------------------------------------------
    # Query: execute SQL and return results
    # -----------------------------------------------------------------

    def execute_sql(
        self,
        collection_name: str,
        sql: str,
        max_rows: int = 50,
    ) -> Dict[str, Any]:
        """Execute a SQL query and return results."""
        db_path = self._db_path(collection_name)
        if not os.path.exists(db_path):
            return {"error": f"No SQL database for collection '{collection_name}'", "rows": [], "columns": []}

        conn = self._connect(collection_name)
        try:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            # Safety: only allow SELECT and PRAGMA
            sql_stripped = sql.strip().upper()
            if not (sql_stripped.startswith("SELECT") or sql_stripped.startswith("PRAGMA")):
                return {"error": "Only SELECT queries are allowed", "rows": [], "columns": []}

            cur.execute(sql)
            columns = [desc[0] for desc in cur.description] if cur.description else []
            rows = [list(row) for row in cur.fetchmany(max_rows)]

            return {
                "columns": columns,
                "rows": rows,
                "row_count": len(rows),
                "sql": sql,
            }
        except Exception as exc:
            return {"error": str(exc), "rows": [], "columns": [], "sql": sql}
        finally:
            conn.close()

    def get_schema(self, collection_name: str) -> str:
        """Return a text description of all tables and columns in the database."""
        db_path = self._db_path(collection_name)
        if not os.path.exists(db_path):
            return "(no structured data tables)"

        conn = self._connect(collection_name)
        try:
            cur = conn.cursor()

            # Get table registry
            try:
                cur.execute("SELECT table_name, source_file, columns, row_count FROM _table_registry")
                registry = cur.fetchall()
            except sqlite3.OperationalError:
                registry = []

            if not registry:
                # Fallback: use sqlite_master
                cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE '_%'")
                tables = cur.fetchall()
                parts = []
                for (tname,) in tables:
                    cur.execute(f'PRAGMA table_info("{tname}")')
                    cols = cur.fetchall()
                    col_names = [c[1] for c in cols]
                    parts.append(f"Table: {tname}\n  Columns: {', '.join(col_names)}")
                return "\n\n".join(parts) if parts else "(no tables)"

            parts = []
            for table_name, source_file, columns, row_count in registry:
                parts.append(
                    f"Table: {table_name}\n"
                    f"  Source: {source_file}\n"
                    f"  Columns: {columns}\n"
                    f"  Rows: {row_count}"
                )

            # Also show sample data for each table
            for table_name, _, _, _ in registry:
                try:
                    cur.execute(f'SELECT * FROM "{table_name}" LIMIT 3')
                    sample_rows = cur.fetchall()
                    if sample_rows:
                        cols = [desc[0] for desc in cur.description]
                        sample_text = f"\n  Sample data from {table_name}:\n"
                        sample_text += f"  | {' | '.join(cols)} |\n"
                        for row in sample_rows:
                            vals = [str(v)[:30] if v else "" for v in row]
                            sample_text += f"  | {' | '.join(vals)} |\n"
                        parts.append(sample_text)
                except Exception:
                    pass

            return "\n\n".join(parts)
        finally:
            conn.close()

    def list_tables(self, collection_name: str) -> List[Dict[str, Any]]:
        """List all tables in the collection's SQL database."""
        db_path = self._db_path(collection_name)
        if not os.path.exists(db_path):
            return []

        conn = self._connect(collection_name)
        try:
            cur = conn.cursor()
            try:
                cur.execute("SELECT table_name, source_file, columns, row_count FROM _table_registry")
                return [
                    {"table_name": r[0], "source_file": r[1], "columns": r[2], "rows": r[3]}
                    for r in cur.fetchall()
                ]
            except sqlite3.OperationalError:
                return []
        finally:
            conn.close()

    def delete_source(self, collection_name: str, source_file: str) -> int:
        """Delete all tables originating from a source file."""
        db_path = self._db_path(collection_name)
        if not os.path.exists(db_path):
            return 0

        conn = self._connect(collection_name)
        deleted = 0
        try:
            cur = conn.cursor()
            try:
                cur.execute(
                    "SELECT table_name FROM _table_registry WHERE source_file = ?",
                    (source_file,),
                )
                tables = cur.fetchall()
                for (tname,) in tables:
                    cur.execute(f'DROP TABLE IF EXISTS "{tname}"')
                    deleted += 1
                cur.execute("DELETE FROM _table_registry WHERE source_file = ?", (source_file,))
                conn.commit()
            except sqlite3.OperationalError:
                pass
        finally:
            conn.close()
        return deleted

    def delete_collection(self, collection_name: str) -> None:
        """Remove the entire SQL database for a collection."""
        db_path = self._db_path(collection_name)
        if os.path.exists(db_path):
            os.remove(db_path)
